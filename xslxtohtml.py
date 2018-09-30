
import openpyxl
from openpyxl import load_workbook
aplh = {-6: '<', -5: '0', -4: '9', -3: '$', -2 : '#', -1: '@', 0: '!', 1: 'a', 2: 'b', 3: 'c', 4: 'd',
        5: 'e', 6: 'f', 7: 'g', 8: 'h', 9: 'i', 10: 'j', 11: 'k',12: 'l', 13: 'm', 14: 'n', 15: 'o', 16: 'p', 17: 'q', 18: 'r', 19:
        's', 20: 't', 21: 'u', 22: 'v', 23: 'w', 24: 'x', 25: 'y', 26: 'z', 27: ' ', 28: '%', 29: '^', 30: '&', 31: '*',
        32: '1', 33: '2', 34: '3', 35: '4', 36: '5', 37: '6', 38: '7', 39: '8', -7: '>', -8: '}', -9: '?', 40: ':'}

def convertxslx(filename):
    workbook = load_workbook(filename)
    worksheet = workbook.get_active_sheet()

    html_data = """
    <html>
        <head>
            <link rel='stylesheet' src= {{ url_for('static', filename = 'main.css') }}>
            <title>
            XLSX to HTML demo
            </title>
        <head>
        <body>
        <table class='table table-hover' border='3'>
    """
    max_row = (worksheet.max_row)
    max_col = (aplh[worksheet.max_column]).upper()
    print(max_col)
    ws_range = worksheet['A1': '{}{}'.format(max_col, max_row)]
    for row in ws_range:
        html_data += "<tr class='table-active'>"
        for cell in row:
            if cell.value is None:
                html_data += "<td> &nbsp </td>"
            else:
                html_data += "<td>" + str(cell.value) + "</td>"
        html_data += "</tr>"
    html_data += "</table></body></html>"

    file = open('static/results.html', 'w+')
    file.write(html_data)
    file.close()
